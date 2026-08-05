"""
文件 → 规范字段提取服务

把上传的业务文件（箱单/发票/舱单等）提取为结构化字段 dict：
- Excel(.xlsx/.xlsm)：openpyxl 按列名匹配 + 按目标柜号列定位数据行（确定性、可测）
- 其他(PDF/图片/docx/txt)：extract_text() 提文本 → LLM extract_json() 按 schema 提取（best-effort）

返回 dict 只含规范 key（见 CANONICAL_KEYS），键缺失即表示该字段未提取到。
"""
from __future__ import annotations

import asyncio
import os
import re

# 规范字段 key（与 merge_service / 佰信模板 source 对齐）
CANONICAL_KEYS = [
    "container_no", "booking_no", "bl_no", "size_type", "seal",
    "gross", "pieces", "volume", "cargo_name",
    "vessel", "voyage", "terminal", "pol", "dest", "etd", "eta",
]

# 数值型 key（提取时转数值，空/0 不写）
NUMERIC_KEYS = {"pieces", "gross", "volume"}

# Excel 列名关键词表（单元格文本规范化后子串匹配，每 key 内按长度降序优先）
COLUMN_KEYWORDS = {
    "container_no": ["柜号/箱号", "集装箱号", "柜号", "箱号", "container no", "container_no",
                     "container", "cn", "ctn"],
    "booking_no": ["订舱单号", "订舱号", "so no", "so号", "booking_no", "booking", "so", "订舱"],
    "cargo_name": ["货物名称", "货名", "品名", "description", "commodity", "cargo", "货物", "goods"],
    "size_type": ["尺寸/类型", "箱型", "柜型", "尺寸", "container type", "size", "type"],
    "pieces": ["件数", "箱数", "pcs", "piece", "quantity", "数量", "件"],
    "gross": ["gross weight", "grossweight", "毛重", "gross", "kg", "gw"],
    "volume": ["材积", "体积", "cbm", "volume", "vol", "方"],
    "seal": ["封条", "封铅", "铅封", "sealno", "seal"],
}

_CN_RE = re.compile(r"[^\d]")


def _norm_cell(v) -> str:
    """单元格文本规范化：转 str、去空白、小写。"""
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v).lower())


def _to_float(v):
    """宽松转 float：剥单位(KGS/CBM/逗号等)，失败返回 None。"""
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.]", "", str(v))
    if not s or s in (".", ""):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize(fields: dict, source: str) -> dict:
    """LLM 输出规范化：只保留规范 key，类型强转，空值/0 剔除。"""
    out = {}
    for k in CANONICAL_KEYS:
        v = fields.get(k)
        if v is None:
            continue
        if isinstance(v, (int, float)):
            s = str(v)
        else:
            s = str(v).strip()
        if not s:
            continue
        if k == "pieces":
            m = re.search(r"\d+", s)
            if m and int(m.group()) > 0:
                out[k] = int(m.group())
        elif k in ("gross", "volume"):
            f = _to_float(s)
            if f is not None and f > 0:
                out[k] = f
        elif k == "container_no":
            s = re.sub(r"[^A-Za-z0-9]", "", s).upper()
            if s:
                out[k] = s
        else:
            out[k] = s
    out["source"] = source
    return out


def _find_header_and_colmap(sheet) -> tuple[dict, int, list[str]]:
    """找表头行并建立 列idx→字段key 映射。返回 (colmap, header_row, warnings)。"""
    max_row = min(sheet.max_row, 15)
    best = None  # (得分, 行号, colmap)
    for r in range(1, max_row + 1):
        row = sheet[r]
        matched = {}  # key -> col idx
        for ci, cell in enumerate(row):
            text = _norm_cell(cell.value)
            if not text or ci >= sheet.max_column:
                continue
            for key, kws in COLUMN_KEYWORDS.items():
                if key in matched:
                    continue
                for kw in kws:
                    if kw in text:
                        matched[key] = ci
                        break
        score = len(matched)
        if score >= 2 and (best is None or score > best[0]):
            best = (score, r, matched)
    if not best:
        return {}, 0, ["未找到表头行（无法识别列名）"]
    return best[2], best[1], []


def _match_data_row(sheet, header_row, colmap, container_no: str) -> tuple[int, list[str]]:
    """在表头行下方找数据行。优先按柜号列匹配目标柜号，否则退回首个非空数据行。"""
    warnings = []
    target = re.sub(r"[^A-Za-z0-9]", "", (container_no or "").upper())
    for r in range(header_row + 1, sheet.max_row + 1):
        row_vals = [c.value for c in sheet[r]]
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue
        if target:
            cidx = colmap.get("container_no")
            if cidx is not None and cidx < len(row_vals):
                cell = _norm_cell(row_vals[cidx]).upper()
                if cell and (target in cell or cell in target):
                    return r, warnings
        # 无 target 或该行柜号不匹配 → 首个非空行作为回退候选
        if not target and r == header_row + 1:
            return r, warnings
    if target:
        warnings.append(f"未在文件中找到柜号 {container_no} 对应数据行，退回首行数据")
        for r in range(header_row + 1, sheet.max_row + 1):
            if any(v is not None and str(v).strip() for v in [c.value for c in sheet[r]]):
                return r, warnings
    return 0, warnings


def _extract_excel_columns(file_path: str, container_no: str = "") -> tuple[dict, list[str]]:
    """Excel 结构化提取：表头行 → 列映射 → 按柜号匹配数据行 → 提取字段。"""
    import openpyxl

    fields, warnings = {}, []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        return fields, [f"Excel 读取失败: {e}"]

    try:
        for ws in wb.worksheets:
            colmap, header_row, ws_warn = _find_header_and_colmap(ws)
            warnings.extend(ws_warn)
            if not colmap:
                continue
            data_row, dr_warn = _match_data_row(ws, header_row, colmap, container_no)
            warnings.extend(dr_warn)
            if not data_row:
                continue
            row = [c.value for c in ws[data_row]]
            for key, ci in colmap.items():
                if ci >= len(row):
                    continue
                val = row[ci]
                if key in NUMERIC_KEYS:
                    if key == "pieces":
                        m = re.search(r"\d+", _norm_cell(val) if val is not None else "")
                        if m and int(m.group()) > 0:
                            fields[key] = int(m.group())
                    else:
                        f = _to_float(val)
                        if f is not None and f > 0:
                            fields[key] = f
                else:
                    s = str(val).strip() if val is not None else ""
                    if s:
                        if key == "container_no":
                            s = re.sub(r"[^A-Za-z0-9]", "", s).upper()
                        fields[key] = s
            if fields:
                break  # 已从第一个可用 sheet 提取到字段
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if not fields:
        warnings.append("Excel 中未提取到有效字段")
    fields["source"] = "excel"
    return fields, warnings


async def _extract_llm_fields(file_path: str, container_no: str,
                              db=None) -> tuple[dict, list[str]]:
    """非 Excel → 文本 → LLM 结构化提取（best-effort，失败返回空 dict）。"""
    from backend.parser import extract_text
    from backend.addons.llm import llm_service

    try:
        text = extract_text(file_path)
    except Exception as e:
        return {}, [f"文本提取失败: {e}"]
    if not text or len(text.strip()) < 5:
        return {}, ["文件文本为空或格式不支持"]

    schema = (
        '{"container_no":"集装箱号 4大写字母+7数字 如ECMU6262406",'
        '"size_type":"箱型 如20GP/40HQ",'
        '"seal":"封条/铅封号",'
        '"gross":"毛重数字KG",'
        '"pieces":"件数整数",'
        '"volume":"体积数字CBM",'
        '"cargo_name":"品名",'
        '"booking_no":"订舱号/SO号",'
        '"bl_no":"提单号",'
        '"vessel":"船名",'
        '"voyage":"航次",'
        '"terminal":"码头",'
        '"pol":"装货港",'
        '"dest":"目的港",'
        '"etd":"ETD日期",'
        '"eta":"ETA日期"}'
    )
    try:
        result = await llm_service.extract_json(
            f"目标柜号:{container_no}\n请从以下文档文本提取字段，缺失的字段留空字符串。\n---\n{text[:15000]}",
            schema, db)
    except Exception as e:
        return {}, [f"LLM 提取失败: {e}"]

    if not isinstance(result, dict) or "error" in result or "raw" in result:
        return {}, ["LLM 提取未返回有效 JSON"]
    fields = normalize(result, "llm")
    if not fields or fields.keys() == {"source"}:
        return {}, ["LLM 未提取到有效字段"]
    return fields, []


async def extract_fields_from_file(file_path: str, container_no: str = "",
                                   db=None) -> tuple[dict, list[str]]:
    """入口：按扩展名分发，返回 (fields, warnings)。fields 含 source 标记。"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return await asyncio.to_thread(_extract_excel_columns, file_path, container_no)
    return await _extract_llm_fields(file_path, container_no, db)
